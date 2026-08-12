"""Kompakter Gesamt-Report-Export (PDF + Excel) — vereinheitlicht.

Eine zentrale Stelle, wo ALLES zusammenläuft:
  • bestehende Parameter (HRV Zeit/vagal/nichtlinear/Frequenz Welch, EEG Bandpower abs+rel,
    Alpha-Gipfel/PAR, Ratios, Verlangsamung/Aperiodik/Komplexität, Asymmetrie),
  • je Wert eine **Korrigiert**-Spalte (artefaktkorrigiert, Auto- oder übergebene Maske),
  • eine Sektion **Validierte Zusatzverfahren** (eigen vs. validiert: Hamilton-R-Zacken, FOOOF,
    Lomb-Scargle, relative Asymmetrie, aperiodik-bereinigter Alpha-Peak, Permutationsentropie).

Je Zeile Wert · Einheit · Norm/Richtungsdeutung. Ausgabe Excel (openpyxl) + PDF (reportlab,
Unicode-Font DejaVu → α₁, ≤, ↑↓, − korrekt). Verändert die bestehenden Analyse-Seiten NICHT.
"""

from __future__ import annotations

import io
import math
import os

import numpy as np

_BK = ["Delta (1–4 Hz)", "Theta (4–8 Hz)", "Alpha (8–13 Hz)", "Beta (13–30 Hz)"]
_BN = ["Delta", "Theta", "Alpha", "Beta"]


def _f(v, fmt=".1f"):
    try:
        if v is None or (isinstance(v, float) and math.isnan(v)):
            return "—"
        return format(float(v), fmt)
    except (TypeError, ValueError):
        return "—"


# ──────────────────────────────────────────────────────────────────────────────
# Signal-/Metrik-Helfer
# ──────────────────────────────────────────────────────────────────────────────
def _clean_sig(sig, fs, segments):
    keep = np.ones(len(sig), dtype=bool)
    for s in segments:
        keep[max(0, int(s["start_s"] * fs)):min(len(sig), int(s["end_s"] * fs))] = False
    return sig[keep]


def _ai(l, r):
    s = l + r
    return (l - r) / s * 100 if s > 1e-9 else float("nan")


def _eeg_metrics(edf, edf_path, segments=None, window_hint_segments=None):
    """Alle EEG-Spektral-Kennzahlen als flaches Dict — einmal Gesamt (segments=None),
    einmal artefaktkorrigiert (segments=Liste, Signale sample-bereinigt)."""
    from views.report import _compute_bandpower
    from analysis.spectral import _highpass, _spectral_edge, _peak_freq_cog
    corrected = bool(segments)
    sf, dur, em = edf["sfreq"], edf["duration_s"], edf["eeg_map"]

    def raw(ch):
        if ch not in em:
            return None
        s = _highpass(edf["data"][em[ch]] * 1e6, sf, 1.0)
        return _clean_sig(s, sf, segments) if corrected else s

    o1, o2, f3, f4 = raw("O1"), raw("O2"), raw("F3"), raw("F4")
    post = (o1 + o2) / 2 if o1 is not None and o2 is not None else (o1 if o1 is not None else o2)
    ant = (f3 + f4) / 2 if f3 is not None and f4 is not None else (f3 if f3 is not None else f4)
    m = {}
    if post is None:
        return m
    if corrected:
        t0, t1 = 0.0, None
    else:
        # Bewusst NICHT ein blindes Mitte-5-min-Fenster (das war der bisherige Ansatz hier
        # und wich vom Visual/Glory-Report ab — Cross-Report-Konsistenzcheck 2026-08-09 fand
        # dadurch abweichende PAR-Werte für dieselbe Aufnahme). Stattdessen dasselbe "beste
        # Alpha-Fenster" wie im Visual Report (analysis/glory_report.py::_best_alpha_window)
        # — sauberes, repräsentatives Fenster statt eines willkürlichen Zeitausschnitts, siehe
        # [[feedback_edf_analysefenster_konsistenz]] und [[project_edf_report_audit]].
        # window_hint_segments: die auto-erkannten Artefakt-Segmente NUR zur Fensterwahl
        # (auch im "Gesamt"-Modus, wo `segments` selbst None ist/nichts aus der Berechnung
        # entfernt wird) — vermeidet, ein artefaktbelastetes Fenster als "repräsentativ"
        # auszuwählen, exakt wie im Visual Report.
        from analysis.glory_report import _best_alpha_window
        _bt, _wl = _best_alpha_window(post, sf, dur, window_hint_segments or [])
        if _bt is not None:
            t0, t1 = _bt, _bt + _wl
        else:
            ana = min(dur, 300.0); t0 = max(0.0, (dur - ana) / 2); t1 = t0 + ana

    bp_p, fp, pp, ap_p = _compute_bandpower(post, sf, t0, t1)
    if not bp_p:
        return m
    res_a = _compute_bandpower(ant, sf, t0, t1) if ant is not None else (None,)
    bp_a = res_a[0] if res_a and res_a[0] else {}
    ap_a = res_a[3] if res_a and res_a[0] else float("nan")
    tp = sum(bp_p.values()) or 1
    ta = sum(bp_a.values()) or 1

    m["rel_post"] = {bn: bp_p.get(bk, 0) / tp * 100 for bk, bn in zip(_BK, _BN)}
    m["abs_post"] = {bn: bp_p.get(bk, 0) for bk, bn in zip(_BK, _BN)}
    m["rel_ant"] = {bn: bp_a.get(bk, 0) / ta * 100 for bk, bn in zip(_BK, _BN)} if bp_a else {}
    m["abs_ant"] = {bn: bp_a.get(bk, 0) for bk, bn in zip(_BK, _BN)} if bp_a else {}
    m["ap_post"], m["ap_ant"] = ap_p, ap_a
    m["cog_post"] = _peak_freq_cog(fp, pp, 8.0, 13.0)
    a_p = bp_p.get("Alpha (8–13 Hz)", 0)
    a_a = bp_a.get("Alpha (8–13 Hz)", 0) if bp_a else 0
    m["ap_ratio"] = a_p / (a_a or 1e-9) if bp_a else float("nan")
    d = bp_p.get("Delta (1–4 Hz)", 0); t = bp_p.get("Theta (4–8 Hz)", 0)
    a = a_p or 1e-9; b = bp_p.get("Beta (13–30 Hz)", 0) or 1e-9
    m["dar"], m["tar"] = d / a, t / a
    m["atr"], m["tbr"], m["dtab"] = a / (t or 1e-9), t / b, (d + t) / (a + b)

    m["sef95"] = _spectral_edge(fp, pp, 0.95)
    m["medf"] = _spectral_edge(fp, pp, 0.50)
    try:
        from analysis.aperiodic import fit_aperiodic, band_power_defs
        from analysis.complexity import sample_entropy, lziv_complexity, permutation_entropy
        rap = fit_aperiodic(fp, pp, 1, 20)
        m["exp_own"] = rap["exponent"] if rap else float("nan")
        m["r2_own"] = rap["r2"] if rap else float("nan")
        m["flat_alpha"] = band_power_defs(fp, pp, 8, 13, res=rap)["flattened"]
        seg = post if corrected else post[int(t0 * sf):int(t1 * sf)]
        m["sampen"] = sample_entropy(seg, max_n=4000) if len(seg) >= 100 else float("nan")
        m["permen"] = permutation_entropy(seg) if len(seg) >= 100 else float("nan")
        m["lzc"] = lziv_complexity(seg, sf) if len(seg) >= int(5 * sf) else {"shuffle": float("nan"), "phase": float("nan")}
    except Exception:
        m["lzc"] = {"shuffle": float("nan"), "phase": float("nan")}

    m["par"], m["exp_grad"] = float("nan"), float("nan")
    if not corrected:
        try:
            from views.eeg_spectrum import _compute_par
            par = _compute_par(edf_path, t0, t1, 8.0, 13.0, False, 9999.0)
            if par["n_post"] >= 2 and par["n_ant"] >= 2:
                m["par"], m["exp_grad"] = par["par"], par["exp_grad"]
        except Exception:
            pass

    m["ai"] = {}
    for lbl, lch, rch in [("O1/O2", "O1", "O2"), ("F3/F4", "F3", "F4")]:
        sl, sr = raw(lch), raw(rch)
        if sl is None or sr is None:
            continue
        bl = _compute_bandpower(sl, sf, t0, t1)[0]
        br = _compute_bandpower(sr, sf, t0, t1)[0]
        if not bl or not br:
            continue
        # tot_l/tot_r statt tl/tr: `tr` ist appweit die Übersetzungsfunktion (core/i18n.py)
        tot_l, tot_r = (sum(bl.values()) or 1), (sum(br.values()) or 1)
        for bk, bn in zip(_BK, _BN):
            m["ai"][(lbl, bn, "abs")] = _ai(bl.get(bk, 0), br.get(bk, 0))
            m["ai"][(lbl, bn, "rel")] = _ai(bl.get(bk, 0) / tot_l, br.get(bk, 0) / tot_r)
    return m


def _compute_hrv_corrected(edf_path, edf, segments):
    from analysis.ecg import detect_r_peaks_polarity_safe, build_rr_series, compute_hrv_time_domain, dfa_alpha1
    from analysis.complexity import sample_entropy
    from analysis.hrv_freq import compute_frequency_domain
    ch = edf["ecg_channels"][0]
    if ch not in edf.get("ch_idx", {}):
        return None
    fs = edf["sfreq"]
    sig = edf["data"][edf["ch_idx"][ch]].astype(float)
    # Polaritäts-sicherer Pfad (User-Audit 2026-08-08) — siehe [[project_edf_rhythm_screening]]
    _, peaks, _ = detect_r_peaks_polarity_safe(sig, fs)
    rr = build_rr_series(peaks, fs)
    if rr is None:
        return None
    rr_ms, times, ect = rr.rr_ms, rr.rr_times_s, rr.artifact_mask
    in_seg = np.zeros(len(rr_ms), dtype=bool)
    for s in segments:
        in_seg |= (times >= s["start_s"]) & (times < s["end_s"])
    keep = (~ect) & (~in_seg)
    rr_c, times_c = rr_ms[keep], times[keep]
    if len(rr_c) < 10:
        return None
    td = compute_hrv_time_domain(rr_c)
    _dfa = dfa_alpha1(rr_c)
    fd = None
    try:
        fd = compute_frequency_domain(rr_c, times_c, method="welch")
    except Exception:
        pass
    return {"mean_hr": td["mean_hr_bpm"], "mean_rr": td["mean_rr_ms"], "sdnn": td["sdnn_ms"],
            "cv": td["cv_pct"], "rmssd": td["rmssd_ms"], "pnn50": td["pnn50_pct"],
            "pnn20": td["pnn20_pct"], "nn50": td["nn50_count"], "sd1": td["sd1_ms"],
            "sd2": td["sd2_ms"], "sd2_sd1": td["sd2_sd1_ratio"],
            "dfa_a1": _dfa["alpha1"] if _dfa else float("nan"),
            "samp_en": sample_entropy(rr_c) if len(rr_c) >= 20 else float("nan"),
            "edr_rate": float("nan"), "pct_removed": float("nan"),
            "fd_welch": fd, "fd_burg": None}


def _hrv_hamilton(edf):
    """HRV-Zeitbereich mit validiertem R-Zacken-Detektor (Hamilton 2002).

    Gibt `None` zurück, wenn Hamilton NICHT lief (Bibliothek fehlt o. Ä.). Bewusst kein
    stiller Rückfall auf den eigenen Detektor: die Werte stehen im Report der Spalte
    „Standard (eigen)" als *Vergleich* gegenüber — käme derselbe Detektor zweimal, zeigte die
    Tabelle zwei identische Spalten und behauptete trotzdem einen Methodenvergleich."""
    from analysis.ecg import (detect_r_peaks_validated_ex, detect_polarity_flip,
                              build_rr_series, compute_hrv_time_domain)
    ch = edf["ecg_channels"][0]
    if ch not in edf.get("ch_idx", {}):
        return None
    fs = edf["sfreq"]
    sig = edf["data"][edf["ch_idx"][ch]].astype(float)
    # Polaritäts-sicherer Pfad (User-Audit 2026-08-08): erst flippen, DANACH den validierten
    # Detektor laufen lassen — sonst verschiebt dessen interne Verfeinerung den Zeitindex bei
    # invertiertem Kanal, siehe [[project_edf_rhythm_screening]].
    if detect_polarity_flip(sig - sig.mean(), fs):
        sig = -sig
    res = detect_r_peaks_validated_ex(sig, fs, "hamilton")
    if res.fell_back:
        return None
    rr = build_rr_series(res.peaks, fs)
    if rr is None:
        return None
    return compute_hrv_time_domain(rr.rr_ms[~rr.artifact_mask])


# ──────────────────────────────────────────────────────────────────────────────
# Sektionen zusammenstellen  →  [{name, columns, rows}]
# ──────────────────────────────────────────────────────────────────────────────
def collect_sections(edf: dict, edf_path: str, corr_segments=None,
                     age=None, sex=None, is_pediatric=False):
    """`age`/`sex`/`is_pediatric`: Patientenkontext für die alters-/HF-adjustierte
    Laborwert-Bewertung (Hansen 2024 / Gąsior 2018, siehe `analysis/report_metadata.py`).
    Fehlt `age`, fällt die Bewertung auf den Erwachsenen-Default (50 J.) zurück — wie
    `core.shared.get_patient_info()` es bei fehlender Eingabe ebenfalls tut."""
    # Haftungshinweis als ERSTE Sektion: ein Report wird ausgedruckt, weitergereicht und
    # in eine Akte gelegt — losgelöst von der App, die ihn erzeugt hat. Was dort nicht
    # draufsteht, steht für den Leser nicht zur Verfügung.
    sections = [{
        "name": "Hinweis",
        "columns": ["Angabe", "Wert"],
        "rows": [["Status", "Kein Medizinprodukt, keine Diagnosesoftware"],
                 ["Zweck", "Forschung, methodische Exploration und Lehre"],
                 ["Werte", "Orientierung — keine Diagnosekriterien, kein Ersatz für die "
                           "ärztliche Befundung"]],
        "wrap": True,
        "col_widths_mm": [30, 156],
    }]
    sf, dur, em = edf["sfreq"], edf["duration_s"], edf.get("eeg_map", {})
    has_ecg = bool(edf.get("ecg_channels"))
    age = age if age is not None else 50

    # Artefakt-Maske für die Korrigiert-Spalte (übergeben oder automatisch)
    if corr_segments is None:
        try:
            from analysis.artifacts import mask_from_edf
            corr_segments = mask_from_edf(edf).segments
        except Exception:
            corr_segments = []
    disc = sum(s["end_s"] - s["start_s"] for s in corr_segments) if corr_segments else 0.0

    # ── Aufnahme ──────────────────────────────────────────────────────────────
    phi = "PHI im Header" if (edf.get("has_patient_id") or edf.get("has_rec_id")) else "anonymisiert"
    sections.append({"name": "Aufnahme & Erkennung",
                     "columns": ["Parameter", "Wert", "Einheit", "Hinweis"], "rows": [
        ["Dauer", f"{dur/60:.1f}", "min", f"{int(dur)} s"],
        ["Abtastrate", _f(sf, ".0f"), "Hz", ""],
        ["Kanäle gesamt", str(len(edf["ch_names"])), "", ""],
        ["EEG-Kanäle (10-20)", str(len(em)), "", ""],
        ["EKG erkannt", "ja" if has_ecg else "nein", "", edf["ecg_channels"][0] if has_ecg else ""],
        ["Datenschutz", phi, "", ""],
        ["Artefakt-Korrektur", f"{len(corr_segments)} Segmente", "",
         f"{disc:.0f}s entfernt · {max(0.0, dur-disc)/dur*100:.0f}% sauber" if dur else ""],
    ]})

    # ── HRV ───────────────────────────────────────────────────────────────────
    # Bewertung alters-/HF-adjustiert (Hansen 2024 / Gąsior 2018), IDENTISCH zur Live-App-
    # Seite views/ecg_hrv.py — vorher zeigte dieser Report statische Erwachsenen-Fixwerte im
    # Text, unabhängig vom tatsächlichen Patientenalter (Report-Audit-Fund, siehe
    # [[project_edf_report_audit]]). "Bewertung"-Spalte ersetzt die alte "Norm / Deutung"-
    # Freitextspalte; Zeilen tragen zusätzlich eine Zone (row["zones"][i]) für die farbliche
    # Hervorhebung in PDF/Excel (siehe build_pdf/build_excel).
    if has_ecg:
        try:
            from views.report import _compute_hrv
            hf = _compute_hrv(edf_path, edf)
        except Exception:
            hf = None
        hc = _compute_hrv_corrected(edf_path, edf, corr_segments) if (hf and corr_segments) else None
        if hf:
            from analysis.report_metadata import grade_hrv, HRV_PARAM_DEFS
            hr_val = hf.get("mean_hr", 70.0) or 70.0
            rmssd_val = hf.get("rmssd")
            _used_params = []

            def _row(param, fmt=".1f", src=None, src_c=None):
                """Baut eine Report-Zeile [Label, Gesamt, Korrigiert, Bewertung, Einheit,
                Referenz] + liefert die Zone separat für die Farbcodierung. Bewertung/
                Referenz beziehen sich auf den GESAMT-Wert (die Korrigiert-Spalte bleibt
                ein reiner Vergleichswert ohne eigene Zweitbewertung, wie schon bisher).
                `src`/`src_c` optional: alternative Quell-Dicts (für die Frequenzbereich-
                Zeilen, deren Werte in hf["fd_welch"] statt direkt in hf liegen)."""
                meta = HRV_PARAM_DEFS[param]
                _key = "mean_hr" if param == "heart_rate" else param  # classify_parameter()
                # nennt den Parameter "heart_rate", _compute_hrv() liefert ihn als "mean_hr"
                _src, _src_c = (src if src is not None else hf), (src_c if src_c is not None else hc)
                gv, cv_ = _src.get(_key), (_src_c.get(_key) if _src_c else None)
                _used_params.append(param)
                extra = {"rmssd_ms": rmssd_val} if param == "pnn50" else {}
                grade = grade_hrv(param, gv, age, hr_val, is_pediatric=is_pediatric, **extra)
                row = [meta["label"], _f(gv, fmt), _f(cv_, fmt) if _src_c else "—",
                       grade["label"], meta["unit"], grade["ref_text"]]
                return row, grade["zone"]

            gc = ["Parameter", "Gesamt", "Korrigiert", "Bewertung", "Einheit", "Referenz"]

            rows, zones = [], []
            for p, fmt in [("heart_rate", ".1f"), ("mean_rr", ".0f"), ("sdnn", ".1f"), ("cv", ".1f")]:
                r, z = _row(p, fmt); rows.append(r); zones.append(z)
            sections.append({"name": "HRV — Zeitbereich", "columns": gc, "rows": rows, "zones": zones})

            rows, zones = [], []
            for p, fmt in [("rmssd", ".1f"), ("pnn50", ".1f"), ("pnn20", ".1f"), ("nn50", ".0f")]:
                r, z = _row(p, fmt); rows.append(r); zones.append(z)
            sections.append({"name": "HRV — vagale Marker", "columns": gc, "rows": rows, "zones": zones})

            rows, zones = [], []
            for p, fmt in [("sd1", ".1f"), ("sd2", ".1f"), ("sd2_sd1", ".2f"), ("dfa_a1", ".2f"),
                           ("samp_en", ".2f"), ("edr_rate", ".1f"), ("pct_removed", ".1f")]:
                r, z = _row(p, fmt); rows.append(r); zones.append(z)
            sections.append({"name": "HRV — nichtlinear & Atmung", "columns": gc, "rows": rows, "zones": zones})

            fw, fwc = (hf.get("fd_welch") or {}), (hc.get("fd_welch") if hc else {}) or {}
            rows, zones = [], []
            for p, fmt in [("total_power", ".0f"), ("vlf_power", ".0f"), ("lf_power", ".0f"),
                           ("hf_power", ".0f"), ("lf_hf_ratio", ".2f"), ("lf_norm", ".1f"),
                           ("hf_norm", ".1f"), ("lf_peak_freq", ".3f"), ("hf_peak_freq", ".3f"),
                           ("hf_resp_rate", ".1f")]:
                r, z = _row(p, fmt, src=fw, src_c=fwc); rows.append(r); zones.append(z)
            sections.append({"name": "HRV — Frequenzbereich (Welch, Task Force 1996)",
                             "columns": gc, "rows": rows, "zones": zones})

            # Begriffserklärungen (Akronyme/Fachbegriffe) — kompakte Liste statt einer 7.
            # Tabellenspalte, damit die Werte-Tabellen selbst schlank/lesbar bleiben
            # (User-Vorgabe: „einfacher Report soll simpel bleiben, aber alles verständlich").
            sections.append({"name": "HRV — Begriffserklärungen", "columns": ["Parameter", "Erklärung"],
                             "rows": [[HRV_PARAM_DEFS[p]["label"], HRV_PARAM_DEFS[p]["definition"]]
                                      for p in dict.fromkeys(_used_params)]})

    # ── EEG ───────────────────────────────────────────────────────────────────
    if em:
        ef = _eeg_metrics(edf, edf_path, None, window_hint_segments=corr_segments)
        ec = _eeg_metrics(edf, edf_path, corr_segments) if corr_segments else {}
        if ef:
            _add_eeg_sections(sections, ef, ec, age=age)

    # ── Validierte Zusatzverfahren (eigen vs. validiert) ──────────────────────
    _add_validated(sections, edf, edf_path, has_ecg, em)

    # ── Herkunft ──────────────────────────────────────────────────────────────
    # Bewusst am ENDE und in beiden Ausgabeformaten: Wer zwei Reports derselben Aufnahme
    # vergleicht, muss entscheiden können, ob ein Unterschied aus der Aufnahme oder aus einer
    # Codeänderung stammt. Ohne diese Angaben ging das nicht.
    _add_provenance(sections, edf, edf_path, corr_segments, age, is_pediatric)
    return sections


def _add_provenance(sections, edf, edf_path, corr_segments, age, is_pediatric):
    from core.version import provenance, provenance_lines
    # Die Parameter, die das Ergebnis tatsächlich verschieben — nicht jede Einstellung der
    # Oberfläche, sondern das, was in die Zahlen eingeht.
    from analysis.spectral import FREQ_MAX
    params = {
        "Artefaktmaske": (f"{len(corr_segments)} Segment(e)" if corr_segments
                          else "keine"),
        "Alter": age,
        "pädiatrische Normwerte": "ja" if is_pediatric else "nein",
        "Abtastrate": f"{edf['sfreq']:g} Hz",
        # Die Spektralparameter gehören in den Report, nicht nur in docs/PREPROCESSING.md:
        # ein Report soll für sich allein nachrechenbar sein. Ohne Fensterlänge, Fenstertyp
        # und Überlapp lässt sich eine Bandpower nicht reproduzieren, auch wenn Version und
        # Commit danebenstehen.
        "EEG-Hochpass": "1 Hz, Butterworth 4. Ordnung, nullphasig (filtfilt)",
        "PSD-Verfahren": f"Welch, Epochen 4 s, Hann, 50 % Überlapp, Density, 1–{FREQ_MAX:g} Hz",
        "Multitaper (optional)": "DPSS, NW=3, K=5",
        "QRS-Detektion": "Bandpass 5–15 Hz (Ordnung 2), 150-ms-Integration, ±40-ms-Refinement",
        "HRV-Frequenzdomäne": "PCHIP-Resampling 4 Hz; Welch und Burg (Ordnung 16)",
        "Analysefenster": "bestes sauberes 60-s-Fenster nach posteriorer Alpha-Relativpower",
    }
    try:
        prov = provenance(edf_path, params)
        rows = [[k, v] for k, v in provenance_lines(prov)]
    except Exception as exc:                      # nie den ganzen Report daran scheitern lassen
        rows = [["Herkunft nicht ermittelbar", str(exc)]]
    sections.append({
        "name": "Herkunft & Reproduzierbarkeit",
        "columns": ["Angabe", "Wert"],
        "rows": rows,
        # Die Paketliste ist lang und die Bezeichnungen sind es teilweise auch — diese
        # Sektion MUSS umbrechen duerfen, sonst laeuft sie aus der Seite.
        "wrap": True,
        "col_widths_mm": [52, 134],
    })


def _add_eeg_sections(sections, ef, ec, age=None):
    from analysis.report_metadata import grade_eeg, EEG_PARAM_DEFS
    gc = ["Parameter", "Gesamt", "Korrigiert", "Einheit", "Norm / Deutung"]
    gc6 = ["Parameter", "Gesamt", "Korrigiert", "Bewertung", "Einheit", "Referenz"]

    def pair(dfull, dcorr, key, sub=None, fmt=".1f"):
        gv = (dfull.get(key, {}).get(sub) if sub else dfull.get(key)) if dfull else None
        cv = (dcorr.get(key, {}).get(sub) if sub else dcorr.get(key)) if dcorr else None
        return _f(gv, fmt), _f(cv, fmt)

    def graded_row(param, key, fmt=".2f"):
        """Baut eine [Label,Gesamt,Korrigiert,Bewertung,Einheit,Referenz]-Zeile — Bewertung
        bezieht sich auf den GESAMT-Wert (wie im HRV-Teil). Nutzt dieselben Zonen-Schwellen
        wie die Live-App (views/eeg_spectrum.py), damit Report und App nicht auseinanderlaufen
        — siehe [[project_edf_report_audit]] für den Gesamtplan der Report-Überarbeitung."""
        meta = EEG_PARAM_DEFS[param]
        gv, cv_ = ef.get(key), (ec.get(key) if ec else None)
        grade = grade_eeg(param, gv, age=age)
        return [meta["label"], _f(gv, fmt), _f(cv_, fmt) if ec else "—",
                grade["label"], meta["unit"], grade["ref_text"]], grade["zone"]

    rows = []
    for bn in _BN:
        g, c = pair(ef, ec, "rel_post", bn)
        rows.append([f"{bn} relativ", g, c, "%", "rel. Anteil (posterior O1/O2)"])
    for bn in _BN:
        g, c = pair(ef, ec, "abs_post", bn, ".2f")
        rows.append([f"{bn} absolut", g, c, "µV²", "posterior"])
    sections.append({"name": "EEG-Bandpower posterior O1/O2", "columns": gc, "rows": rows})

    if ef.get("rel_ant"):
        rows = []
        for bn in _BN:
            g, c = pair(ef, ec, "rel_ant", bn)
            note = "↑ Delta anterior oft EOG-Artefakt" if bn == "Delta" else "anterior"
            rows.append([f"{bn} relativ", g, c, "%", note])
        sections.append({"name": "EEG-Bandpower anterior F3/F4", "columns": gc, "rows": rows})

    rows6, zones6 = [], []
    for param, key, fmt in [("ap_post", "ap_post", ".2f"), ("ap_ratio", "ap_ratio", ".2f"),
                            ("par", "par", ".2f")]:
        r, z = graded_row(param, key, fmt); rows6.append(r); zones6.append(z)
    # Ungraded Zusatzwerte (kein etablierter eigener Cutoff) im selben, größeren Rahmen
    rows6.append(["Alpha-Peak CoG posterior", *pair(ef, ec, "cog_post", fmt=".2f"), "—", "Hz",
                 "Schwerpunkt (robuster als Gipfel)"])
    zones6.append("info")
    rows6.append(["Alpha-Gipfel anterior", *pair(ef, ec, "ap_ant", fmt=".2f"), "—", "Hz",
                 "erwartet < posterior"])
    zones6.append("info")
    rows6.append(["Exponent-Gradient post−ant", _f(ef.get("exp_grad"), "+.2f"), "—", "—", "—",
                 "+ = posterior steiler (nur Gesamt)"])
    zones6.append("info")
    sections.append({"name": "EEG — Alpha-Gipfel & A/P-Gradient", "columns": gc6,
                     "rows": rows6, "zones": zones6})

    # Klinische Frequenzratios: bewusst OHNE Bewertungs-Spalte — die App selbst nutzt hierfür
    # keine festen, klinisch validierten Cutoffs (Texte sind explizit "orientierend"/
    # "Frühmarker"), eine erfundene Ampel-Schwelle wäre hier Overclaiming.
    sections.append({"name": "EEG — klinische Frequenzratios (posterior)", "columns": gc, "rows": [
        ["Delta/Alpha (DAR)", *pair(ef, ec, "dar", fmt=".3f"), "Ratio", "0–1,5 · ↑ Verlangsamung (orientierend)"],
        ["Theta/Alpha (TAR)", *pair(ef, ec, "tar", fmt=".3f"), "Ratio", "0,2–0,7 · Frühmarker (orientierend)"],
        ["Alpha/Theta", *pair(ef, ec, "atr", fmt=".3f"), "Ratio", "1,5–6 · Vigilanz (orientierend, ↑ = wach)"],
        ["Theta/Beta (TBR)", *pair(ef, ec, "tbr", fmt=".3f"), "Ratio", "0,5–2 · Schläfrigkeit (orientierend)"],
        ["DTAB (D+T)/(A+B)", *pair(ef, ec, "dtab", fmt=".3f"), "Ratio", "< 0,5 · kort. Funktion (orientierend)"],
    ]})

    lzf = ef.get("lzc", {}); lzc = ec.get("lzc", {}) if ec else {}
    rows6, zones6 = [], []
    r, z = graded_row("exp_own", "exp_own", ".2f")
    r[5] += f" · R²={_f(ef.get('r2_own'), '.2f')}"
    rows6.append(r); zones6.append(z)
    for lbl, key, fmt, note in [
        ("SEF95", "sef95", ".1f", "↓ = Verlangsamung (orientierend)"),
        ("Medianfrequenz (SEF50)", "medf", ".1f", "↓ = Verlangsamung (orientierend)"),
        ("Alpha flattened", "flat_alpha", ".2f", "> 0 = echter Gipfel über dem 1/f-Untergrund"),
        ("Sample Entropy", "sampen", ".2f", "↓ = regelmäßig (kein etablierter Cutoff)"),
        ("Permutationsentropie", "permen", ".2f", "Bandt-Pompe · ↓ = regelmäßig (kein Cutoff)"),
    ]:
        g, c = pair(ef, ec, key, fmt=fmt)
        unit = "Hz" if key in ("sef95", "medf") else "—"
        rows6.append([lbl, g, c, "—", unit, note])
        zones6.append("info")
    rows6.append(["LZC (shuffle)", _f(lzf.get("shuffle"), ".2f"), _f(lzc.get("shuffle"), ".2f"),
                 "—", "—", "↑ = komplex (kein etablierter Cutoff)"])
    zones6.append("info")
    rows6.append(["LZC (phase)", _f(lzf.get("phase"), ".2f"), _f(lzc.get("phase"), ".2f"),
                 "—", "—", "> 1 = spektral-unabhängig"])
    zones6.append("info")
    sections.append({"name": "EEG — Verlangsamung, Aperiodik (1/f) & Komplexität",
                     "columns": gc6, "rows": rows6, "zones": zones6})

    # Asymmetrie: Gesamt(abs) + Korrigiert(abs); relative Variante in Validiert-Sektion
    rows6, zones6 = [], []
    for lbl in ("O1/O2", "F3/F4"):
        for bn in _BN:
            gv = ef.get("ai", {}).get((lbl, bn, "abs"))
            cv = ec.get("ai", {}).get((lbl, bn, "abs")) if ec else None
            grade = grade_eeg("ai", gv)
            rows6.append([f"AI {bn} ({lbl})", _f(gv, ".0f"), _f(cv, ".0f"),
                         grade["label"], "%", grade["ref_text"]])
            zones6.append(grade["zone"])
    sections.append({"name": "EEG — Hemisphärische Asymmetrie (absolut)", "columns": gc6,
                     "rows": rows6, "zones": zones6})

    sections.append({"name": "EEG — Begriffserklärungen", "columns": ["Parameter", "Erklärung"],
                     "rows": [[v["label"], v["definition"]] for v in EEG_PARAM_DEFS.values()]})


def _add_validated(sections, edf, edf_path, has_ecg, em):
    gc = ["Parameter", "Standard (eigen)", "Validiert", "Einheit", "Referenz / Hinweis"]
    rows = []
    # R-Zacken: eigen vs Hamilton (HRV-Zeit)
    if has_ecg:
        try:
            from views.report import _compute_hrv
            hf = _compute_hrv(edf_path, edf)
        except Exception:
            hf = None
        ham = _hrv_hamilton(edf)
        if hf and ham:
            rows += [
                ["SDNN (R-Zacken-Detektor)", _f(hf.get("sdnn")), _f(ham.get("sdnn_ms")), "ms", "Hamilton 2002 (py-ecg-detectors)"],
                ["RMSSD (R-Zacken-Detektor)", _f(hf.get("rmssd")), _f(ham.get("rmssd_ms")), "ms", "sensibel für Timing-Präzision"],
                ["pNN50 (R-Zacken-Detektor)", _f(hf.get("pnn50")), _f(ham.get("pnn50_pct")), "%", "Hamilton"],
            ]
        elif hf:
            # Grund nennen statt die Zeilen kommentarlos wegzulassen — sonst wirkt ein
            # unvollständiger Report wie ein vollständiger.
            from analysis.ecg import validated_detectors_available
            _why = ("py-ecg-detectors nicht installiert"
                    if not validated_detectors_available()
                    else "validierter Detektor lieferte kein verwertbares Ergebnis")
            rows.append(["R-Zacken-Detektor (Vergleich)", _f(hf.get("sdnn")), "—", "ms",
                         f"kein Vergleich möglich: {_why}"])
        # DFA: eigen (nicht überlappend, nur α1) vs Standard-DFA (überlappend, α1+α2)
        try:
            from views.ecg_hrv import compute_rr as _crr
            from analysis.ecg import dfa_alpha12
            _rrm = _crr(edf_path, edf["ecg_channels"][0])["rr_ms"]
            a12 = dfa_alpha12(_rrm)
            if a12 and hf:
                rows += [
                    ["DFA α1", _f(hf.get("dfa_a1"), ".2f"), _f(a12["alpha1"], ".2f"), "—",
                     "eigen (nicht überlappend) vs Standard-DFA (überlappend)"],
                    ["DFA α2 (16–64 Schläge)", "—", _f(a12["alpha2"], ".2f"), "—",
                     "Langzeit-Steigung — nur im Standard-DFA (neu)"],
                ]
        except Exception:
            pass
        # HRV-Spektrum: Welch vs Lomb-Scargle
        try:
            from views.ecg_hrv import compute_rr
            from analysis.hrv_freq import compute_frequency_domain
            from analysis.hrv_lombscargle import lombscargle_hrv
            rr = compute_rr(edf_path, edf["ecg_channels"][0])
            w = compute_frequency_domain(rr["rr_ms"], rr["times"], method="welch")
            ls = lombscargle_hrv(rr["rr_ms"], rr["times"])
            if w and ls:
                rows += [
                    ["LF/HF-Ratio (HRV-Spektrum)", _f(w.get("lf_hf_ratio"), ".2f"), _f(ls.get("lf_hf_ratio"), ".2f"), "Ratio", "Lomb-Scargle: interpolationsfrei"],
                    ["LF normiert (HRV-Spektrum)", _f(w.get("lf_norm")), _f(ls.get("lf_norm")), "%", "Welch vs Lomb-Scargle"],
                    ["HF normiert (HRV-Spektrum)", _f(w.get("hf_norm")), _f(ls.get("hf_norm")), "%", "Welch vs Lomb-Scargle"],
                ]
        except Exception:
            pass
    # Aperiodik: eigen vs FOOOF; Alpha-Peak CoG vs FOOOF
    if em:
        try:
            from analysis.spectral import _highpass, _peak_freq_cog
            from analysis.aperiodic import welch_psd, fit_aperiodic
            from analysis.aperiodic_fooof import fit_fooof
            sf, dur = edf["sfreq"], edf["duration_s"]
            ch = "O2" if "O2" in em else ("O1" if "O1" in em else list(em)[0])
            ana = min(dur, 300.0); t0 = max(0.0, (dur - ana) / 2)
            sig = _highpass(edf["data"][em[ch]] * 1e6, sf, 1.0)
            f, p = welch_psd(sig[int(t0 * sf):int((t0 + ana) * sf)], sf, fmax=45.0)
            own = fit_aperiodic(f, p, 1, 40)
            ff = fit_fooof(f, p, 1, 40, knee=False)
            if own and ff:
                rows += [
                    [f"Aperiod. Exponent 1–40 Hz ({ch})", _f(own["exponent"], ".2f"), _f(ff["exponent"], ".2f"), "—", f"FOOOF R²={_f(ff['r2'], '.2f')} vs eigen R²={_f(own['r2'], '.2f')}"],
                ]
                own_a = _peak_freq_cog(f, p, 8, 13)
                fa = [pk for pk in ff["peaks"] if 8 <= pk[0] <= 13]
                fa_v = max(fa, key=lambda x: x[1])[0] if fa else float("nan")
                rows += [[f"Alpha-Peak ({ch})", _f(own_a, ".2f"), _f(fa_v, ".2f"), "Hz", "eigen CoG (linear) vs FOOOF (aperiodik-bereinigt)"]]
        except Exception:
            pass
        # relative Asymmetrie (validiert gegen absolute)
        try:
            ef = _eeg_metrics(edf, edf_path, None)
            for lbl in ("O1/O2", "F3/F4"):
                for bn in _BN:
                    a_abs = ef.get("ai", {}).get((lbl, bn, "abs"))
                    a_rel = ef.get("ai", {}).get((lbl, bn, "rel"))
                    if a_abs is None or a_rel is None:
                        continue
                    rows.append([f"AI {bn} ({lbl})", _f(a_abs, ".0f"), _f(a_rel, ".0f"), "%", "absolut vs relativ (impedanz-robust)"])
        except Exception:
            pass
    if rows:
        sections.append({"name": "Validierte Zusatzverfahren (eigen vs. validiert)",
                         "columns": gc, "rows": rows})


# ──────────────────────────────────────────────────────────────────────────────
# Excel
# ──────────────────────────────────────────────────────────────────────────────
_ZONE_FILL = {"normal": "E3F2E3", "grenzwertig": "FDF3D9", "pathologisch": "FBE1DE"}
_ZONE_FONT = {"normal": "1E7B34", "grenzwertig": "9C6F00", "pathologisch": "B23A24"}


def build_excel(sections, edf, disp_name: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws.append([f"EDF-Analyzer — Gesamt-Report · {disp_name}"])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([])
    for sec in sections:
        r = ws.max_row + 1
        ws.cell(row=r, column=1, value=sec["name"]).font = Font(bold=True, color="1F4E79")
        ws.append(sec["columns"])
        for c in ws[ws.max_row]:
            c.font = Font(bold=True)
        # "Bewertung"-Spalte farblich hervorheben (Laborwert-Stil: grün/gelb/rot), sofern
        # die Sektion Zonen mitliefert (nur HRV-Wertetabellen — siehe collect_sections()).
        zones = sec.get("zones")
        bew_col = sec["columns"].index("Bewertung") + 1 if "Bewertung" in sec["columns"] else None
        for i, row in enumerate(sec["rows"]):
            ws.append(list(row))
            if zones and bew_col and i < len(zones) and zones[i] in _ZONE_FILL:
                cell = ws.cell(row=ws.max_row, column=bew_col)
                cell.fill = PatternFill("solid", fgColor=_ZONE_FILL[zones[i]])
                cell.font = Font(color=_ZONE_FONT[zones[i]], bold=True)
        ws.append([])
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 12), 52)

    # Kanäle + Ereignisse
    ch = wb.create_sheet("Kanäle")
    ch.append(["Nr", "Kanal", "Min", "Max", "RMS", "Einheit"])
    for i, name in enumerate(edf["ch_names"]):
        sig = edf["data"][i] - edf["data"][i].mean()
        unit = "µV" if name.startswith("EEG") else "mV"
        fac = 1e6 if name.startswith("EEG") else 1e3
        ch.append([i, name, round(float(sig.min() * fac), 1), round(float(sig.max() * fac), 1),
                   round(float(np.sqrt(np.mean(sig ** 2)) * fac), 1), unit])
    if edf.get("annotations"):
        ev = wb.create_sheet("Ereignisse")
        ev.append(["Zeit (s)", "Ereignis"])
        for a in edf["annotations"]:
            ev.append([round(a["onset_s"], 1), a["description"]])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ──────────────────────────────────────────────────────────────────────────────
# PDF (DejaVu-Font → volle Unicode-Unterstützung)
# ──────────────────────────────────────────────────────────────────────────────
def _register_font():
    try:
        import matplotlib
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        base = os.path.join(matplotlib.get_data_path(), "fonts", "ttf")
        pdfmetrics.registerFont(TTFont("RepSans", os.path.join(base, "DejaVuSans.ttf")))
        pdfmetrics.registerFont(TTFont("RepSans-Bold", os.path.join(base, "DejaVuSans-Bold.ttf")))
        return "RepSans", "RepSans-Bold"
    except Exception:
        return "Helvetica", "Helvetica-Bold"


def build_pdf(sections, disp_name: str) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    font, font_b = _register_font()
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=13 * mm, bottomMargin=12 * mm,
                            leftMargin=12 * mm, rightMargin=12 * mm, title="EDF-Report")
    styles = getSampleStyleSheet()
    title = ParagraphStyle("t", parent=styles["Title"], fontName=font_b)
    normal = ParagraphStyle("n", parent=styles["Normal"], fontName=font)
    h_sec = ParagraphStyle("sec", parent=styles["Heading4"], fontName=font_b, spaceBefore=7,
                           spaceAfter=2, textColor=colors.HexColor("#2471a3"))
    story = [Paragraph("EDF-Analyzer — Gesamt-Report", title),
             Paragraph(f"Datei: {disp_name}", normal), Spacer(1, 5)]
    # Spaltenbreiten je nach Spaltenzahl (2/4/5/6)
    widths = {2: [40 * mm, 146 * mm],
              4: [60 * mm, 40 * mm, 22 * mm, 64 * mm],
              5: [56 * mm, 27 * mm, 27 * mm, 18 * mm, 58 * mm],
              6: [40 * mm, 20 * mm, 20 * mm, 28 * mm, 26 * mm, 52 * mm]}   # Bewertung breiter
    _zone_bg = {"normal": colors.HexColor("#e3f2e3"), "grenzwertig": colors.HexColor("#fdf3d9"),
               "pathologisch": colors.HexColor("#fbe1de")}
    _zone_fg = {"normal": colors.HexColor("#1e7b34"), "grenzwertig": colors.HexColor("#9c6f00"),
               "pathologisch": colors.HexColor("#b23a24")}
    # Zellenstil fuer umbrechende Sektionen. Reportlab bricht rohe Strings NICHT um: ein zu
    # langer Wert laeuft stumm ueber den Satzspiegel hinaus oder ueberdruckt die Nachbarzelle.
    # Aufgefallen beim ersten Ansehen des Herkunft-Abschnitts — die Paketliste lief rechts aus
    # der Seite, "pädiatrische Normwerte" ueberschrieb seinen eigenen Wert.
    cell = ParagraphStyle("cell", parent=normal, fontSize=7.5, leading=9)
    # Zahlen rechtsbündig — sonst lassen sich Gesamt- und Korrigiert-Spalte nicht mehr
    # untereinander vergleichen. Die Table-ALIGN-Regeln greifen bei Paragraph-Zellen nicht
    # mehr, die Ausrichtung muss deshalb in den Absatzstil.
    cell_r = ParagraphStyle("cell_r", parent=cell, alignment=2)   # 2 = rechts
    # Welche Spalten Zahlen tragen, hängt an der Spaltenzahl — dieselbe Zuordnung, die die
    # bisherigen ALIGN-Regeln weiter unten verwenden.
    _num_cols = {4: (1, 2), 5: (1, 2, 3), 6: (1, 2, 3)}
    for sec in sections:
        cols = sec["columns"]
        story.append(Paragraph(sec["name"], h_sec))
        rows = [list(r) for r in sec["rows"]]
        if sec.get("wrap"):
            rows = [[Paragraph(str(c), cell) for c in r] for r in rows]
        else:
            # Rohe Strings brechen in ReportLab NICHT um: eine zu lange Bewertung wie
            # „leicht-mäßig grenzwertig" lief in die Nachbarzelle und überdruckte deren Wert
            # (im erzeugten PDF gesehen, in keinem Test). Deshalb brechen jetzt ALLE
            # Sektionen um — der Preis sind etwas höhere Zeilen, der Gewinn ist, dass nichts
            # mehr stillschweigend unlesbar wird.
            num = _num_cols.get(len(cols), ())
            rows = [[Paragraph(str(c), cell_r if i in num else cell)
                     for i, c in enumerate(r)] for r in rows]
        data = [cols] + rows
        tbl = Table(data, colWidths=sec.get("col_widths_mm") and
                    [w * mm for w in sec["col_widths_mm"]] or widths.get(len(cols)),
                    repeatRows=1)
        style = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eef3fb")),
            ("FONTNAME", (0, 0), (-1, -1), font),
            ("FONTNAME", (0, 0), (-1, 0), font_b),
            ("FONTSIZE", (0, 0), (-1, -1), 7.5),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f7f9fc")]),
            ("LINEBELOW", (0, 0), (-1, 0), 0.6, colors.HexColor("#c4ccd6")),
            ("TOPPADDING", (0, 0), (-1, -1), 2), ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]
        if len(cols) in (5, 6):
            style.append(("ALIGN", (1, 0), (3, -1), "RIGHT"))
        elif len(cols) == 4:
            style.append(("ALIGN", (1, 0), (2, -1), "RIGHT"))
        # "Bewertung"-Spalte farblich hervorheben (Laborwert-Stil), analog build_excel()
        zones = sec.get("zones")
        if zones and "Bewertung" in cols:
            bcol = cols.index("Bewertung")
            for i, z in enumerate(zones):
                if z in _zone_bg:
                    r = i + 1  # +1: Header-Zeile
                    style.append(("BACKGROUND", (bcol, r), (bcol, r), _zone_bg[z]))
                    style.append(("TEXTCOLOR", (bcol, r), (bcol, r), _zone_fg[z]))
                    style.append(("FONTNAME", (bcol, r), (bcol, r), font_b))
        tbl.setStyle(TableStyle(style))
        story.append(tbl)
    doc.build(story)
    return buf.getvalue()


# ──────────────────────────────────────────────────────────────────────────────
# Maschinenlesbares Manifest
# ──────────────────────────────────────────────────────────────────────────────
def build_manifest(sections, edf: dict, edf_path: str, disp_name: str,
                   age=None, sex=None, is_pediatric: bool = False) -> bytes:
    """Dasselbe wie der Report, nur für Maschinen: JSON statt Tabelle.

    PDF und Excel sind für Menschen gemacht. Wer zwei Aufnahmen vergleichen, über eine Serie
    auswerten oder ein Ergebnis in einer anderen Umgebung nachrechnen will, muss sie derzeit
    abtippen. Dieses Manifest enthält dieselben Werte samt vollständiger Herkunft in einer
    Form, die sich einlesen und diffen lässt.

    Bewusst OHNE Rohdaten und ohne Kopfdaten der Aufnahme: die Datei wird über ihren SHA-256
    identifiziert, nicht über Namen oder Patientenfelder. Ein Manifest kann damit
    weitergegeben werden, auch wenn die Aufnahme es nicht darf.
    """
    import json
    from core.version import provenance

    # Sektionen in eine flache, stabil benannte Struktur bringen. Die Reihenfolge der Zeilen
    # bleibt erhalten, damit ein Diff zweier Manifeste lesbar bleibt.
    inhalte = []
    for sec in sections:
        inhalte.append({
            "section": sec["name"],
            "columns": list(sec["columns"]),
            "rows": [[_jsonfaehig(c) for c in row] for row in sec["rows"]],
        })

    prov = provenance(edf_path, {
        "Alter": age, "Geschlecht": sex,
        "pädiatrische Normwerte": bool(is_pediatric),
        "Abtastrate_Hz": float(edf.get("sfreq", 0)),
        "Dauer_s": round(float(edf.get("duration_s", 0)), 3),
    })

    manifest = {
        # Versionierung des FORMATS, nicht der App — wer das Manifest einliest, muss wissen,
        # ob sich die Struktur geändert hat, unabhängig von der Analyse-Version.
        "manifest_schema": "1.0",
        "tool": "EDF-Analyzer",
        "created_utc": _jetzt_utc(),
        "provenance": prov,
        "recording": {
            # Kein Dateiname, keine Patientenfelder — siehe Docstring.
            "sha256": prov.get("file_sha256"),
            "duration_s": round(float(edf.get("duration_s", 0)), 3),
            "sfreq_hz": float(edf.get("sfreq", 0)),
            "n_eeg_channels": len(edf.get("eeg_map", {})),
            "ecg_channels": list(edf.get("ecg_channels") or []),
        },
        "results": inhalte,
    }
    return json.dumps(manifest, ensure_ascii=False, indent=2).encode("utf-8")


def _jetzt_utc() -> str:
    from datetime import datetime, timezone
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _jsonfaehig(wert):
    """NumPy-Typen und NaN in etwas verwandeln, das `json` schreiben kann.

    `json.dumps` schreibt für NaN das Literal `NaN` — das ist **kein gültiges JSON** und
    lässt jeden strengen Parser scheitern. Fehlende Werte werden deshalb zu `null`.
    """
    if isinstance(wert, (str, bool)) or wert is None:
        return wert
    if isinstance(wert, (int, float)):
        return None if isinstance(wert, float) and math.isnan(wert) else wert
    if isinstance(wert, np.generic):          # np.float64, np.int64 …
        w = wert.item()
        return None if isinstance(w, float) and math.isnan(w) else w
    return str(wert)
